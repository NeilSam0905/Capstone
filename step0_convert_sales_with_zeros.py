"""
Step 0 (new): re-convert the USTore TBS tally-sheet workbooks into long
format, this time PRESERVING true zero-sale days instead of dropping them.

Why this exists
----------------
The original converter ("Converter Aug 2024 - May 2026.py") drops any cell
that is blank OR an explicit 0 - it can't tell the two apart, so it treats
every gap in an item's series as "not observed". That's correct for a
sheet that only has a handful of tally-date columns (a periodic physical
stock count), but wrong for a sheet that has a date column for nearly
every calendar day in the month - there, a blank cell means "the store
was tallied that day and this item sold zero", a real data point that
was being silently discarded everywhere downstream (Fact_Sales, ADUS,
Prophet's series).

Per-sheet check: (# of real date columns) / (days spanned by those
columns). Verified against every monthly sheet in the source workbooks:
  - Aug 2024: 5 dates / 22-day span  (0.23) -> sparse, episodic
  - Sep 2024: 4 dates / 19-day span  (0.21) -> sparse, episodic
  - Oct 2024 onward, every month through Jul 2026: consistently >=0.85,
    most at or above 1.0 -> dense, essentially a full daily matrix
The May 2024 DSR&TBS file's "TBS" sheet is also dense (23/30 = 0.77),
consistent with its underlying daily "DAILY SALES REPORT" sheets (May 2,
May 3, ...) which record explicit 0s per item per day.

So: dense sheets get every blank/zero cell filled in as an explicit
Total Quantity = 0 row. Sparse sheets (only Aug/Sep 2024) keep the
original drop-blank-cells behaviour, since a gap there really does mean
"not counted that day", not "counted as zero".

Output: USTore_sales_long_with_zeros.csv (same 4 columns as the
original: Date, Item, Total Quantity, Supplier), reading directly from
the original source workbooks in drive-download-20260724T120738Z-1-001/.

Date is written as ISO 8601 (YYYY-MM-DD). The source workbooks use a
mix of DD/MM/YYYY, MM/DD/YYYY and real Excel date cells in their column
headers; parse_date_header() below resolves those, and everything this
script writes is ISO from that point on. Do not open the output in
Excel - it will silently rewrite the column back to a locale format.
"""
import calendar
import csv
import datetime
import math
import statistics
from collections import defaultdict

import openpyxl

SRC_DIR = "drive-download-20260724T120738Z-1-001"
FILES = [
    f"{SRC_DIR}/2024 5 MAY DSR & TBS.xlsx",
    f"{SRC_DIR}/USTore TBS AUG-DEC 2024.xlsx",
    f"{SRC_DIR}/2025 USTore TBS.xlsx",
    f"{SRC_DIR}/USTore TBS OCTOBER A.Y. 2025-2026.xlsx",
]
OUT_PATH = "USTore_sales_long_with_zeros.csv"
DENSE_THRESHOLD = 0.7

MONTHS = {
    "JAN", "JANUARY", "FEB", "FEBRUARY", "MAR", "MARCH", "APR", "APRIL",
    "MAY", "JUN", "JUNE", "JUL", "JULY", "AUG", "AUGUST", "SEP", "SEPT",
    "SEPTEMBER", "OCT", "OCTOBER", "NOV", "NOVEMBER", "DEC", "DECEMBER",
}
META_HEADERS = {"TOTAL QUANTITY", "ITEM PRICE", "FOR REMITTANCE"}


def is_tbs_month_sheet(name: str) -> bool:
    n = name.strip().upper()
    if not n.endswith("- TBS"):
        return False
    return n.split()[0] in MONTHS


def is_total_row(label) -> bool:
    return isinstance(label, str) and label.strip().upper() == "TOTAL"


def parse_date_header(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        d = value.date() if isinstance(value, datetime.datetime) else value
        return "date", d
    if isinstance(value, str):
        s = value.strip().upper()
        if s == "NO DATE":
            return "nodate", None
        if s in META_HEADERS or s == "":
            return None, None
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y"):
            try:
                return "date", datetime.datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None, None


def round_half_up(x) -> int:
    return int(math.floor(x + 0.5))


def to_quantity(value):
    """Coerce a cell to a quantity, or None if truly blank/non-numeric.
    NOTE: unlike the original converter, this DOES distinguish 0 from
    None - a real 0 is returned as 0.0, not collapsed into None."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def fmt_qty(q):
    return int(q) if float(q).is_integer() else q


def impute_nodate_dates(date_cols, warnings, sheet_title):
    imputed = {}
    if not any(k == "nodate" for _, k, _ in date_cols):
        return imputed
    real = [(c, d) for c, k, d in date_cols if k == "date"]
    if not real:
        warnings.append(f"[{sheet_title}] NO DATE column but no real dates to anchor - left as 'NO DATE'")
        return imputed
    year, month = real[0][1].year, real[0][1].month
    last_day = calendar.monthrange(year, month)[1]
    median_day = round_half_up(statistics.median(sorted(d.day for _, d in real)))
    for i, (c, kind, _) in enumerate(date_cols):
        if kind != "nodate":
            continue
        left = next((date_cols[j][2].day for j in range(i - 1, -1, -1)
                     if date_cols[j][1] == "date"), None)
        right = next((date_cols[j][2].day for j in range(i + 1, len(date_cols))
                      if date_cols[j][1] == "date"), None)
        day = round_half_up((left + right) / 2) if (left is not None and right is not None) else median_day
        day = min(max(day, 1), last_day)
        imputed[c] = datetime.date(year, month, day)
    return imputed


def sheet_density(date_cols):
    """Density = (# real date columns) / (days in that calendar month).
    Uses the FULL month as the denominator, not just the span between the
    earliest and latest found date - a sheet whose only usable dates
    happen to cluster together (e.g. Sep 2024's 17/18/19) would otherwise
    look artificially "dense" just because those few dates are close
    together, even though the month as a whole was barely tallied."""
    real_dates = [d for _, k, d in date_cols if k == "date"]
    if not real_dates:
        return 0.0, 0
    year, month = real_dates[0].year, real_dates[0].month
    span_days = calendar.monthrange(year, month)[1]
    return (len(real_dates) / span_days if span_days else 0.0), span_days


def convert_sheet(ws, sheet_label, warnings, density_log):
    date_cols = []
    relevant_cols = []  # date/meta columns only - excludes stray trailing
                         # annotation columns like "LEGEND" comments
    for c in range(2, ws.max_column + 1):
        header = ws.cell(1, c).value
        kind, d = parse_date_header(header)
        if kind:
            date_cols.append([c, kind, d])
            relevant_cols.append(c)
        elif isinstance(header, str) and header.strip().upper() in META_HEADERS:
            relevant_cols.append(c)
    if not date_cols:
        warnings.append(f"[{sheet_label}] no date columns found - skipped")
        return

    density, span_days = sheet_density(date_cols)
    is_dense = density >= DENSE_THRESHOLD
    n_real = sum(1 for _, k, _ in date_cols if k == "date")
    density_log.append((sheet_label, n_real, span_days, density, is_dense))

    imputed = impute_nodate_dates(date_cols, warnings, sheet_label)

    supplier = None
    prev_was_header_row = False
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None or str(label).strip() == "":
            continue
        label = str(label).strip()

        if is_total_row(label):
            prev_was_header_row = False
            continue

        # A real item row always has SOME data somewhere in its row (a
        # price, a quantity, a remittance figure...). A supplier/
        # consignment label row has NONE - not even a price. This is more
        # robust than only trusting an explicit "TOTAL" row to signal the
        # next label is a new supplier: some sheets are missing that TOTAL
        # row (the label after it then gets a blank cell instead), and a
        # few suppliers split their qualifier onto its own line (e.g.
        # "STITCH CORP. (BLEEVES)" then a separate "(Consignment)" line) -
        # both cases would otherwise get misread as zero-selling "items".
        has_any_data = any(
            ws.cell(r, c).value not in (None, "")
            for c in relevant_cols
        )
        if not has_any_data:
            supplier = f"{supplier} {label}" if prev_was_header_row and supplier else label
            prev_was_header_row = True
            continue
        prev_was_header_row = False

        if supplier is None:
            warnings.append(f"[{sheet_label}] item '{label}' before any supplier - skipped")
            continue

        for c, kind, d in date_cols:
            qty = to_quantity(ws.cell(r, c).value)
            if kind == "date":
                obs = d
            elif c in imputed:
                obs = imputed[c]
            else:
                if qty is not None and qty != 0:
                    yield (1, 0), "NO DATE", label, qty, supplier
                continue

            if qty is None:
                if is_dense:
                    qty = 0.0
                else:
                    continue  # sparse month: a gap really is "not observed"

            yield (0, obs.toordinal()), obs.strftime("%Y-%m-%d"), label, qty, supplier


def convert(files, out_path):
    agg = defaultdict(float)
    meta = {}
    warnings = []
    density_log = []
    per_sheet = []

    for fn in files:
        wb = openpyxl.load_workbook(fn, data_only=True)
        short = fn.split("/")[-1]
        if short == "2024 5 MAY DSR & TBS.xlsx":
            sheet_names = ["TBS"]
        else:
            sheet_names = [sn for sn in wb.sheetnames if is_tbs_month_sheet(sn)]

        for sn in sheet_names:
            label = f"{short} :: {sn}"
            count = 0
            for sort_key, date_str, item, qty, supplier in convert_sheet(wb[sn], label, warnings, density_log):
                key = (date_str, item, supplier)
                agg[key] += qty
                meta[key] = sort_key
                count += 1
            per_sheet.append((short, sn, count))

    rows = sorted(
        ((meta[k], k) for k in agg),
        key=lambda x: (x[0], x[1][2].upper(), x[1][1].upper()),
    )

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Date", "Item", "Total Quantity", "Supplier"])
        for _, (date_str, item, supplier) in rows:
            w.writerow([date_str, item, fmt_qty(agg[(date_str, item, supplier)]), supplier])

    return per_sheet, warnings, density_log, len(agg)


def main():
    per_sheet, warnings, density_log, n = convert(FILES, OUT_PATH)

    print("=== Per-sheet density (real tally dates / calendar-day span) ===")
    for label, n_real, span, density, is_dense in density_log:
        tag = "DENSE (zero-filled)" if is_dense else "sparse (gaps kept as unobserved)"
        print(f"  {label:65} {n_real:3} dates / {span:3} days = {density:.2f}  -> {tag}")

    print("\n=== Per-sheet row counts ===")
    for fname, sheet, c in per_sheet:
        print(f"  {fname:38} | {sheet:22} | {c:6} rows")

    if warnings:
        print("\nNotes / imputations:")
        for wmsg in warnings:
            print("  -", wmsg)

    print(f"\nWrote {n} rows -> {OUT_PATH}")


if __name__ == "__main__":
    main()
