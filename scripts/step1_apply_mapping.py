"""
Step 1 of ETL: apply the canonical-name mapping to the sales and inventory
CSVs, report any unmapped item names, and populate Dim_Product.

Does NOT do proportional allocation and does NOT touch Fact_Sales.
"""
import re
import sqlite3
import sys

import openpyxl
import pandas as pd

from step0_convert_sales_with_zeros import FILES as TBS_FILES, is_tbs_month_sheet

# Remediation S12. Distinct from tools/audit_price_suffix_skus.py's
# PRICE_SUFFIX_RE (r"\s*@.*$"), which strips the suffix to recover the
# base name and throws the number away - this one needs the number
# itself, as a fallback price. "Lanyard @180" -> 180.
PRICE_SUFFIX_RE = re.compile(r"@(\d+)")


def price_from_suffix(item_name):
    m = PRICE_SUFFIX_RE.search(item_name)
    return float(m.group(1)) if m else None


# May 2024's workbook has a "TBS" summary sheet (what step0 reads for
# quantities) AND 23 daily "DAILY SALES REPORT" sheets ("May 2", "May 3", ...)
# that no script has ever touched. Those carry a RETAIL PRICE column - a third
# price fallback, used only for canonical items inventory and the name-suffix
# both missed.
MAY_2024_DSR_WORKBOOK = "rawdata/2024 5 MAY DSR & TBS.xlsx"
MAY_2024_NON_DSR_SHEETS = {"TS", "TBS", "INVENTORY"}


def load_may2024_dsr_prices(mapping):
    """Each daily sheet repeats a header row (ITEMS | RETAIL PRICE | PCS SOLD |
    SALES | DISCOUNTED PRICE | ...) before every supplier's block - the very
    first one has the supplier name on its own row above; every later one has
    the supplier name fused into the header row's first cell instead. A TOTAL
    row (blank item cell, "TOTAL" elsewhere) closes each block. Detecting a
    block boundary by column 2 == "RETAIL PRICE" (rather than by column 1)
    handles both forms without caring which one it is.

    Read-only against data/vocab_mapping_FINAL_v5.csv: a raw name here that
    isn't already in the approved mapping is skipped and counted, never
    guessed at - this is a supplementary price source, not a mapping change.
    """
    wb = openpyxl.load_workbook(MAY_2024_DSR_WORKBOOK, data_only=True)
    sheets = [s for s in wb.sheetnames if s not in MAY_2024_NON_DSR_SHEETS]

    raw_prices = {}  # raw item name -> [retail prices seen across the month]
    for sheet_name in sheets:
        ws = wb[sheet_name]
        in_block = False
        for r in range(1, ws.max_row + 1):
            c1, c2 = ws.cell(r, 1).value, ws.cell(r, 2).value
            if isinstance(c2, str) and c2.strip().upper() == "RETAIL PRICE":
                in_block = True
                continue
            if not in_block:
                continue
            label = str(c1).strip() if c1 is not None else ""
            if not label:
                continue  # TOTAL row, or a stray blank
            if isinstance(c2, (int, float)):
                raw_prices.setdefault(label, []).append(float(c2))

    unmapped = set()
    canonical_prices = {}  # canonical_item_name -> [retail prices seen]
    for raw_name, prices in raw_prices.items():
        canonical = mapping.get(raw_name)
        if canonical is None:
            unmapped.add(raw_name)
            continue
        canonical_prices.setdefault(canonical, []).extend(prices)

    def mode(values):
        return max(set(values), key=values.count)

    result = {canonical: mode(prices) for canonical, prices in canonical_prices.items()}
    n_conflict = sum(1 for prices in canonical_prices.values() if len(set(prices)) > 1)

    print(f"[may2024_dsr] {len(sheets)} daily sheets read, "
          f"{len(raw_prices)} raw item name(s) with a retail price found")
    print(f"[may2024_dsr] {len(unmapped)} raw name(s) not in the vocab mapping - skipped, "
          f"not aborting (supplementary source, not the core mapping)")
    print(f"[may2024_dsr] {len(result)} canonical item(s) priced from this source; "
          f"{n_conflict} had more than one distinct retail price across the month "
          f"(modal value kept)")
    return result


def load_tbs_item_prices(mapping):
    """Every TBS-pattern sheet, in every workbook (not just May 2024's), carries
    an ITEM PRICE column that step0 has never read - it only reads the per-date
    quantity cells. A price sometimes genuinely drifts between months (e.g.
    "UST College ID Lace" ~140 in late 2024, ~100 from 2025 onward) rather than
    being noise, so the modal value across every month an item appears in is
    kept - the same aggregation rule build_dim_product() already uses for
    inv_price. Used only for canonical items inventory, the name-suffix and
    may2024_dsr_price all missed.
    """
    raw_prices = {}
    for fn in TBS_FILES:
        wb = openpyxl.load_workbook(fn, data_only=True)
        short = fn.split("/")[-1]
        if short == "2024 5 MAY DSR & TBS.xlsx":
            sheet_names = ["TBS"]
        else:
            sheet_names = [sn for sn in wb.sheetnames if is_tbs_month_sheet(sn)]

        for sn in sheet_names:
            ws = wb[sn]
            price_col = None
            for c in range(2, ws.max_column + 1):
                header = ws.cell(1, c).value
                if isinstance(header, str) and header.strip().upper() == "ITEM PRICE":
                    price_col = c
                    break
            if price_col is None:
                continue
            for r in range(2, ws.max_row + 1):
                label = ws.cell(r, 1).value
                if label is None:
                    continue
                label = str(label).strip()
                if not label or label.upper() == "TOTAL":
                    continue
                price = ws.cell(r, price_col).value
                if isinstance(price, (int, float)):
                    raw_prices.setdefault(label, []).append(float(price))

    unmapped = set()
    canonical_prices = {}
    for raw_name, prices in raw_prices.items():
        canonical = mapping.get(raw_name)
        if canonical is None:
            unmapped.add(raw_name)
            continue
        canonical_prices.setdefault(canonical, []).extend(prices)

    def mode(values):
        return max(set(values), key=values.count)

    result = {c: mode(p) for c, p in canonical_prices.items()}
    n_conflict = sum(1 for p in canonical_prices.values() if len(set(p)) > 1)

    print(f"[tbs_item_price] {len(raw_prices)} raw item name(s) with a price "
          f"found across every workbook's TBS sheets")
    print(f"[tbs_item_price] {len(unmapped)} raw name(s) not in the vocab mapping - skipped")
    print(f"[tbs_item_price] {len(result)} canonical item(s) priced from this source; "
          f"{n_conflict} had more than one distinct price across the months it "
          f"appeared in (modal value kept - most are real price drift, not noise)")
    return result


MAPPING_CSV = "data/vocab_mapping_FINAL_v5.csv"
SUPPLIER_MAPPING_CSV = "data/supplier_mapping.csv"
SALES_CSV = "data/USTore_sales_long_with_zeros.csv"
INVENTORY_CSV = "data/USTore_inventory_excel_long.csv"
DB_PATH = "ustore.db"

# Named after its actual input (SALES_CSV), not the older pre-zero-fill
# file - proportional_allocation.py reads this, so the name has to say
# which sales file it came from.
SALES_MAPPED_CSV = "data/USTore_sales_long_with_zeros_mapped.csv"
INVENTORY_MAPPED_CSV = "data/USTore_inventory_excel_long_mapped.csv"


def load_mapping():
    vm = pd.read_csv(MAPPING_CSV)
    vm["raw_name"] = vm["raw_name"].astype(str).str.strip()
    vm["canonical_item_name"] = vm["canonical_item_name"].astype(str).str.strip()
    return dict(zip(vm["raw_name"], vm["canonical_item_name"]))


def load_supplier_mapping():
    """42 raw Supplier strings -> 19 suppliers + a payment_status, per
    supplier_mapping.csv. Two of the raw strings resolve to no supplier at
    all (the bare "(Paid)" parser artefact and an item name typed into the
    Supplier column); those carry an empty supplier_name and are counted
    as unattributed rather than invented."""
    sm = pd.read_csv(SUPPLIER_MAPPING_CSV, dtype=str).fillna("")
    for col in ("raw_supplier", "supplier_name", "payment_status"):
        sm[col] = sm[col].str.strip()
    return (dict(zip(sm["raw_supplier"], sm["supplier_name"])),
            dict(zip(sm["raw_supplier"], sm["payment_status"])))


def apply_supplier_mapping(sales, name_map, status_map):
    raw = sales["Supplier"].astype(str).str.strip()
    unknown = sorted(set(raw) - set(name_map))
    if unknown:
        print("[supplier] UNMAPPED supplier strings:")
        for u in unknown:
            print(f"   - {u!r}")
        sys.exit(f"ABORTING: {len(unknown)} supplier string(s) missing from "
                 f"{SUPPLIER_MAPPING_CSV}. Add them there first.")
    sales = sales.copy()
    sales["supplier_name"] = raw.map(name_map).replace("", pd.NA)
    sales["payment_status"] = raw.map(status_map).replace("", pd.NA)
    n_named = int(sales["supplier_name"].notna().sum())
    print(f"[supplier] {raw.nunique()} raw strings -> "
          f"{sales['supplier_name'].nunique()} suppliers; "
          f"{len(sales) - n_named} row(s) with no attributable supplier")
    return sales


def apply_mapping(df, mapping, label):
    stripped_items = df["Item"].astype(str).str.strip()
    canonical = stripped_items.map(mapping)
    unmatched_mask = canonical.isna()
    unmatched_names = sorted(stripped_items[unmatched_mask].unique().tolist())
    df = df.copy()
    df["canonical_item_name"] = canonical
    print(f"[{label}] rows processed: {len(df)}")
    print(f"[{label}] unmatched distinct item names: {len(unmatched_names)}")
    if unmatched_names:
        print(f"[{label}] UNMATCHED NAMES:")
        for name in unmatched_names:
            print(f"   - {name!r}")
    return df, unmatched_names


def build_dim_product(sales_mapped, inventory_mapped, may2024_dsr_price, tbs_item_price):
    # entry_date = earliest sales date per canonical item
    sales_dates = sales_mapped.copy()
    # Every CSV in this repo stores dates as ISO 8601 (YYYY-MM-DD). errors="raise"
    # is deliberate: a coerced date becomes NaT and silently vanishes from the
    # entry_date min(), which nothing downstream would notice.
    sales_dates["parsed_date"] = pd.to_datetime(
        sales_dates["Date"], format="%Y-%m-%d", errors="raise"
    )
    entry_dates = (
        sales_dates.groupby("canonical_item_name")["parsed_date"]
        .min()
        .dt.strftime("%Y-%m-%d")
    )

    # category + unit_price_php: from inventory, most frequent non-null value per item
    def mode_or_none(series):
        s = series.dropna()
        if s.empty:
            return None
        return s.mode().iloc[0]

    inv_cat = inventory_mapped.groupby("canonical_item_name")["Category"].apply(mode_or_none)
    inv_price = inventory_mapped.groupby("canonical_item_name")["Price"].apply(mode_or_none)

    # supplier_name / payment_status: from sales, most frequent non-null value
    # per item, off the NORMALISED columns. Payment status is really a property
    # of a consignment agreement, not of a product, and a few items appear
    # under both terms - the modal value is a summary, and the count of items
    # where it isn't unanimous is reported below so it can't pass unnoticed.
    sales_supplier = sales_mapped.groupby("canonical_item_name")["supplier_name"].apply(mode_or_none)
    sales_status = sales_mapped.groupby("canonical_item_name")["payment_status"].apply(mode_or_none)

    mixed = (
        sales_mapped.dropna(subset=["payment_status"])
        .groupby("canonical_item_name")["payment_status"].nunique()
    )
    n_mixed = int((mixed > 1).sum())
    print(f"[supplier] items sold under more than one payment_status: {n_mixed} "
          f"(Dim_Product keeps the modal value)")

    all_items = sorted(
        set(sales_mapped["canonical_item_name"]) | set(inventory_mapped["canonical_item_name"])
    )

    # Remediation S12. unit_price_php previously had exactly one source -
    # the inventory sheets - and inherited that source's coverage gap
    # wholesale (239 of 519 products, 82.3% of units, unpriced; 48 of 58
    # Fast SKUs). 64 of the 71 price-suffixed products ("Lanyard @180")
    # carry the price in the name itself, the May 2024 workbook's daily
    # sheets carry a RETAIL PRICE column, and every workbook's TBS sheets
    # (every month, not just May) carry their own ITEM PRICE column - none
    # of which any script had read before now. price_source records which
    # of the four supplied the value, so they're never silently conflated -
    # a handful of items where they disagree (name-coined price vs. current
    # inventory price, most likely price drift over time) stay visible
    # rather than being overwritten one way or the other.
    price_source_counts = {
        "inventory": 0, "name_suffix": 0, "may2024_dsr": 0, "tbs_item_price": 0, None: 0,
    }
    rows = []
    for item in all_items:
        price = inv_price.get(item)
        # inv_price is a groupby().apply(mode_or_none) Series: once ANY group
        # returns a real float, pandas coerces the whole Series to float64,
        # silently turning mode_or_none's Python None (no inventory price at
        # all) into np.nan instead. `price is not None` is True for NaN, so
        # this used to mislabel a no-price item as price_source="inventory" -
        # the exact conflation this column exists to prevent. pd.notna()
        # catches both cases.
        if pd.notna(price):
            source = "inventory"
        else:
            price = price_from_suffix(item)
            if price is not None:
                source = "name_suffix"
            else:
                price = may2024_dsr_price.get(item)
                if price is not None:
                    source = "may2024_dsr"
                else:
                    price = tbs_item_price.get(item)
                    source = "tbs_item_price" if price is not None else None
        price_source_counts[source] += 1

        rows.append(
            {
                "item_name": item,
                "category": inv_cat.get(item),
                "unit_price_php": price,
                "price_source": source,
                "supplier_name": sales_supplier.get(item),
                "payment_status": sales_status.get(item),
                "lead_time_days": None,
                "fsn_class": None,
                "entry_date": entry_dates.get(item),
                "is_active": 1,
            }
        )
    print(f"[price] source: inventory={price_source_counts['inventory']}, "
          f"name_suffix={price_source_counts['name_suffix']}, "
          f"may2024_dsr={price_source_counts['may2024_dsr']}, "
          f"tbs_item_price={price_source_counts['tbs_item_price']}, "
          f"unpriced={price_source_counts[None]}")
    return pd.DataFrame(rows)


def main():
    mapping = load_mapping()
    may2024_dsr_price = load_may2024_dsr_prices(mapping)
    tbs_item_price = load_tbs_item_prices(mapping)

    sales = pd.read_csv(SALES_CSV)
    inventory = pd.read_csv(INVENTORY_CSV)

    sales_mapped, sales_unmatched = apply_mapping(sales, mapping, "sales")
    inventory_mapped, inventory_unmatched = apply_mapping(inventory, mapping, "inventory")

    name_map, status_map = load_supplier_mapping()
    sales_mapped = apply_supplier_mapping(sales_mapped, name_map, status_map)

    sales_mapped.to_csv(SALES_MAPPED_CSV, index=False)
    inventory_mapped.to_csv(INVENTORY_MAPPED_CSV, index=False)

    total_unmatched = len(set(sales_unmatched) | set(inventory_unmatched))
    if total_unmatched:
        print(f"\nABORTING Dim_Product load: {total_unmatched} unmatched name(s) found. Fix the mapping file first.")
        sys.exit(1)

    dim_product = build_dim_product(sales_mapped, inventory_mapped, may2024_dsr_price, tbs_item_price)

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("DELETE FROM Dim_Product")
    dim_product.to_sql("Dim_Product", con, if_exists="append", index=False)
    con.commit()
    row_count = cur.execute("SELECT COUNT(*) FROM Dim_Product").fetchone()[0]
    con.close()

    print("\n=== SUMMARY ===")
    print(f"sales rows processed: {len(sales_mapped)}")
    print(f"inventory rows processed: {len(inventory_mapped)}")
    print(f"unmatched names: {total_unmatched}")
    print(f"Dim_Product rows: {row_count}")
    print(f"\nMapped files written: {SALES_MAPPED_CSV}, {INVENTORY_MAPPED_CSV}")


if __name__ == "__main__":
    main()
