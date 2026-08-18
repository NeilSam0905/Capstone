"""
USTore Tally-Sheet (TBS) converter
==================================
Converts the USTore's vertical / wide-format monthly tally-sheet workbooks into a
single tidy (long-format) CSV with the columns:

    Date, Item, Total Quantity, Supplier

Each input .xlsx holds one year of sales; each worksheet holds one month. In every
TBS sheet, row 1 is the header: each column after column A is either a tally DATE,
a "NO DATE" column, or a metadata column (TOTAL QUANTITY / ITEM PRICE / FOR
REMITTANCE). Column A holds supplier section headers, item names, and TOTAL
subtotal rows. Per-date cells hold the units sold of that item on that date.

The script "melts" the date columns into rows: one output row per
(date, item, supplier) with the quantity sold. Blank / zero cells (no sale) are
dropped. Dates are written as ISO 8601 (YYYY-MM-DD) - the source headers use a
mix of DD/MM/YYYY, MM/DD/YYYY and real Excel date cells, and every date this
script emits is normalised to ISO. Do not open the output CSV in Excel; it
rewrites the column back to a locale format.

NO DATE handling
----------------
A "NO DATE" column holds sales that were tallied without a specific day. Rather
than dropping them, each NO DATE column is assigned an imputed date equal to the
MIDPOINT (rounded half-up) of the two recorded tally dates that flank it in the
header order. Example: a month recorded on the 6th, 9th, [NO DATE], 16th, 22nd
puts the NO DATE between the 9th and 16th -> imputed to the 13th. If a NO DATE
column sits at the very start or end (only one neighbour), it falls back to the
median tally day of that month. If a sheet has no real dates at all to anchor to,
the column is left labelled "NO DATE".

Usage:
    python ustore_tbs_to_csv.py                 # uses the three bundled files
    python ustore_tbs_to_csv.py file1.xlsx ...  # or pass your own workbooks
    python ustore_tbs_to_csv.py -o out.csv f.xlsx
"""

import sys
import csv
import math
import argparse
import calendar
import datetime
import statistics
from collections import defaultdict

import openpyxl
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------- #
MONTHS = {
    "JAN", "JANUARY", "FEB", "FEBRUARY", "MAR", "MARCH", "APR", "APRIL",
    "MAY", "JUN", "JUNE", "JUL", "JULY", "AUG", "AUGUST", "SEP", "SEPT",
    "SEPTEMBER", "OCT", "OCTOBER", "NOV", "NOVEMBER", "DEC", "DECEMBER",
}
META_HEADERS = {"TOTAL QUANTITY", "ITEM PRICE", "FOR REMITTANCE"}


def is_tbs_month_sheet(name: str) -> bool:
    """A sales sheet ends in '- TBS' and starts with a real month name.
    This skips 'RE-CHECKING FEB ...', 'VOUCHER PAYMENT ...', 'OPEX', etc."""
    n = name.strip().upper()
    if not n.endswith("- TBS"):
        return False
    return n.split()[0] in MONTHS


def is_total_row(label) -> bool:
    return isinstance(label, str) and label.strip().upper() == "TOTAL"


def parse_date_header(value):
    """Return ('date', date) / ('nodate', None) / (None, None) for a header cell."""
    if isinstance(value, (datetime.datetime, datetime.date)):
        d = value.date() if isinstance(value, datetime.datetime) else value
        return "date", d
    if isinstance(value, str):
        s = value.strip().upper()
        if s == "NO DATE":
            return "nodate", None
        if s in META_HEADERS or s == "":
            return None, None
        # Fallback: try to parse a date typed as text (rare in these files)
        for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y"):
            try:
                return "date", datetime.datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None, None


def round_half_up(x) -> int:
    """13 for 12.5 (Python's built-in round() would give 12 via banker's rounding)."""
    return int(math.floor(x + 0.5))


def to_quantity(value):
    """Coerce a cell to a positive quantity, or None if empty/zero/non-numeric."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value if float(value) != 0 else None
    s = str(value).strip().replace(",", "")
    try:
        f = float(s)
        return f if f != 0 else None
    except ValueError:
        return None  # notes / non-numeric marks are ignored


def fmt_qty(q):
    """Whole numbers -> int (no trailing .0); otherwise keep the number."""
    return int(q) if float(q).is_integer() else q


def impute_nodate_dates(date_cols, warnings, sheet_title):
    """Assign an imputed date to every NO DATE column.

    date_cols: ordered list of [col_index, kind, date_or_None].
    Returns {col_index: datetime.date} for the NO DATE columns that were resolved.
    """
    imputed = {}
    if not any(k == "nodate" for _, k, _ in date_cols):
        return imputed

    real = [(c, d) for c, k, d in date_cols if k == "date"]
    if not real:
        warnings.append(f"[{sheet_title}] NO DATE column but no real dates to anchor - left as 'NO DATE'")
        return imputed

    year, month = real[0][1].year, real[0][1].month
    last_day = calendar.monthrange(year, month)[1]
    median_day = round_half_up(statistics.median(sorted(d.day for _, d in real)))

    for i, (c, kind, _) in enumerate(date_cols):
        if kind != "nodate":
            continue
        left = next((date_cols[j][2].day for j in range(i - 1, -1, -1)
                     if date_cols[j][1] == "date"), None)
        right = next((date_cols[j][2].day for j in range(i + 1, len(date_cols))
                      if date_cols[j][1] == "date"), None)
        if left is not None and right is not None:
            day, how = round_half_up((left + right) / 2), f"midpoint of {left:02d} and {right:02d}"
        else:
            day, how = median_day, "median tally day (edge column)"
        day = min(max(day, 1), last_day)
        d = datetime.date(year, month, day)
        imputed[c] = d
        warnings.append(f"[{sheet_title}] NO DATE -> {d.strftime('%Y-%m-%d')} ({how})")
    return imputed


def convert_sheet(ws, warnings):
    """Yield (sort_key, date_str, item, qty, supplier) tuples for one sheet."""
    # --- ordered date columns from the header (row 1), skipping column A --- #
    date_cols = []  # [col_index, kind, date_or_None]
    for c in range(2, ws.max_column + 1):
        kind, d = parse_date_header(ws.cell(1, c).value)
        if kind:
            date_cols.append([c, kind, d])
    if not date_cols:
        warnings.append(f"[{ws.title}] no date columns found - skipped")
        return

    imputed = impute_nodate_dates(date_cols, warnings, ws.title)

    supplier = None
    expecting_supplier = True  # first non-empty label row is a supplier header
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None or str(label).strip() == "":
            continue
        label = str(label).strip()

        if is_total_row(label):
            expecting_supplier = True   # next label row starts a new supplier block
            continue
        if expecting_supplier:
            supplier = label
            expecting_supplier = False
            continue

        # --- item row: melt its date columns --- #
        if supplier is None:
            warnings.append(f"[{ws.title}] item '{label}' before any supplier - skipped")
            continue
        for c, kind, d in date_cols:
            qty = to_quantity(ws.cell(r, c).value)
            if qty is None:
                continue
            if kind == "date":
                obs = d
            elif c in imputed:      # NO DATE column resolved to an imputed date
                obs = imputed[c]
            else:                   # unresolved NO DATE (no anchor dates)
                yield (1, 0), "NO DATE", label, qty, supplier
                continue
            yield (0, obs.toordinal()), obs.strftime("%Y-%m-%d"), label, qty, supplier


def convert(files, out_path):
    # aggregate by (date, item, supplier); sums any duplicate date columns
    agg = defaultdict(float)
    meta = {}          # key -> sort_key
    warnings = []
    per_sheet = []

    for fn in files:
        wb = openpyxl.load_workbook(fn, data_only=True)
        for sn in wb.sheetnames:
            if not is_tbs_month_sheet(sn):
                continue
            count = 0
            for sort_key, date_str, item, qty, supplier in convert_sheet(wb[sn], warnings):
                key = (date_str, item, supplier)
                agg[key] += qty
                meta[key] = sort_key
                count += 1
            per_sheet.append((fn.split("/")[-1], sn, count))

    rows = sorted(
        ((meta[k], k) for k in agg),
        key=lambda x: (x[0], x[1][2].upper(), x[1][1].upper()),  # date, supplier, item
    )

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Date", "Item", "Total Quantity", "Supplier"])
        for _, (date_str, item, supplier) in rows:
            w.writerow([date_str, item, fmt_qty(agg[(date_str, item, supplier)]), supplier])

    return per_sheet, warnings, len(agg)


def main():
    default_files = [
        "USTore_TBS_AUG-DEC_2024.xlsx",
        "2025_USTore_TBS.xlsx",
        "USTore_TBS_OCTOBER_A_Y__2025-2026.xlsx",
    ]
    ap = argparse.ArgumentParser(description="Convert USTore TBS workbooks to a tidy CSV.")
    ap.add_argument("files", nargs="*", default=default_files, help="Input .xlsx files")
    ap.add_argument("-o", "--output", default="data/USTore_sales_long.csv", help="Output CSV path")
    args = ap.parse_args()

    per_sheet, warnings, n = convert(args.files, args.output)

    print("Per-sheet observation counts:")
    for fname, sheet, c in per_sheet:
        print(f"  {fname:38} | {sheet:22} | {c:5} rows")
    if warnings:
        print("\nNotes / imputations:")
        for wmsg in warnings:
            print("  -", wmsg)
    print(f"\nWrote {n} rows -> {args.output}")


if __name__ == "__main__":
    main()
