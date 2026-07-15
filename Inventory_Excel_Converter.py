"""
USTore Inventory Report (Excel) converter
==========================================
Converts "Copy of USTORE INVENTORY REPORT (1).xlsx" - a 37-sheet workbook
(one MAIN STORAGE snapshot + one APPAREL and one NON-APPAREL sheet per month,
Nov 2024 - Apr 2026) - into a single tidy long-format CSV:

    Category, Date, No, Item, Size, Location, Classification, Price,
    Quantity, OPEX, Notes

Date is a single ISO (YYYY-MM-DD) column: the sheet's exact "DATE: MM/DD/YYYY"
where present (Main Storage, and the Nov/Dec 2024 Apparel sheets), otherwise the
first day of the sheet's "<Month> <Year> INVENTORY" title month.

One row per item / size / location snapshot as of that sheet's report date.
Each monthly Apparel/Non-Apparel sheet also has daily delivery columns
(e.g. "JAN.7", "JAN.8") after the PICTURES column; those are intentionally
NOT melted into rows here (scope: current stock snapshot, not a movement
log) - only the columns up to and including PICTURES/TAKEN ITEMS are read.

Layouts handled (auto-detected per header block; a sheet can contain more
than one block, e.g. Non-Apparel sheets switch to a simpler block partway
down for items that have no classification breakdown):
  - PLAIN : NO, PRICE, ITEM, QUANTITY, [OPEX], [DELIVERY], PICTURES, [TAKEN ITEMS]
  - SIZE  : NO, PRICE, ITEM, SIZE, QUANTITY, PICTURES, [TAKEN ITEMS]
  - CLASS : NO, PRICE, ITEM, CLASSIFICATION, QUANTITY, [OPEX], [DELIVERY], PICTURES
  - LOC   : NO, PRICE, ITEM, INSIDE(SIZE, QUANTITY), OUTSIDE(SIZE, QUANTITY),
            [OPEX], [TOTAL QTY.], PICTURES, <daily dates ignored>

Within a block, a size/classification breakdown for one item spans several
rows: only the first row of the group carries ITEM (and usually NO / PRICE);
the rest are blank there and are forward-filled from the group's first row.

Usage:
    python "Inventory Excel Converter.py"
    python "Inventory Excel Converter.py" file.xlsx -o out.csv
"""

import re
import csv
import datetime
import argparse

import openpyxl

DEFAULT_INPUT = "Copy of USTORE INVENTORY REPORT (1).xlsx"
DEFAULT_OUTPUT = "USTore_inventory_excel_long.csv"

DATE_RE = re.compile(r"DATE:\s*(\d{1,2}/\d{1,2}/\d{4})")
MONTH_RE = re.compile(r"([A-Za-z]+)\.?\s+(\d{4})\s+INVENTORY", re.I)
NUM_RE = re.compile(r"[\d,]+(?:\.\d+)?")


def to_iso_date(report_date, report_month):
    """Consolidate to a single ISO (YYYY-MM-DD) date. Use the sheet's exact
    'DATE: MM/DD/YYYY' where present; otherwise fall back to the first day of the
    sheet's report month, so every row carries a usable date."""
    rd = (report_date or "").strip()
    rm = (report_month or "").strip()
    if rd:
        for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
            try:
                return datetime.datetime.strptime(rd, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    if rm:
        for fmt in ("%B %Y", "%b %Y"):
            try:
                return datetime.datetime.strptime(rm, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return ""


def clean(v):
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))  # Excel stores "No." as a float (1.0); drop the trailing .0
    return str(v).strip()


def find_report_date(rows):
    for row in rows:
        for cell in row:
            if isinstance(cell, str):
                m = DATE_RE.search(cell)
                if m:
                    return m.group(1)
    return ""


def find_report_month(rows, report_date):
    """Prefer the sheet's own 'MONTH YYYY INVENTORY' title; not every sheet has an
    exact DATE:, but every sheet has this title. Fall back to the exact date."""
    for row in rows:
        for cell in row:
            if isinstance(cell, str):
                m = MONTH_RE.search(cell)
                if m:
                    return f"{m.group(1).title()} {m.group(2)}"
    if report_date:
        try:
            return datetime.datetime.strptime(report_date, "%m/%d/%Y").strftime("%B %Y")
        except ValueError:
            pass
    return ""


def category_for(sheet_name):
    """Excel truncates sheet names to 31 chars, so 'FEBRUARY 2025 INVENTORY  FOR AP'
    never actually contains the full word 'APPAREL'. NON-APPAREL sheet names always
    keep the 'NON' substring within the limit, so check that first and treat every
    other non-Main-Storage sheet as Apparel (the only two monthly categories)."""
    n = sheet_name.upper()
    if "MAIN STORAGE" in n:
        return "MAIN STORAGE"
    if "NON" in n:
        return "NON-APPAREL"
    return "APPAREL"


def parse_num(value):
    """Pull a number out of an int/float cell or a text cell like '213 pcs.' / '9,980 pcs.'."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        v = float(value)
        return int(v) if v.is_integer() else v
    m = NUM_RE.search(str(value))
    if not m:
        return None
    f = float(m.group(0).replace(",", ""))
    return int(f) if f.is_integer() else f


def col_of(header_row, *targets):
    """Index of the first cell whose cleaned upper text matches any of targets."""
    for i, c in enumerate(header_row):
        if clean(c).upper().rstrip(".") in targets:
            return i
    return None


def parse_header_block(rows, i):
    """Given rows[i] is a header row (has a PRICE cell), return (schema, first_data_row_idx)."""
    header = rows[i]
    price_col = col_of(header, "PRICE")
    no_col = price_col - 1 if price_col and price_col > 0 else None
    item_col = price_col + 1 if price_col is not None else None

    keywords = {
        "INSIDE": None, "OUTSIDE": None, "CLASSIFICATION": None, "SIZE": None,
        "QUANTITY": None, "OPEX": None, "DELIVERY": None, "PICTURES": None,
        "TAKEN ITEMS": None, "TOTAL QTY": None,
    }
    for idx, c in enumerate(header):
        key = clean(c).upper().rstrip(".")
        if key in keywords and keywords[key] is None:
            keywords[key] = idx

    schema = {"no": no_col, "item": item_col, "price": price_col,
               "opex": keywords["OPEX"], "pictures": keywords["PICTURES"],
               "taken_items": keywords["TAKEN ITEMS"], "delivery": keywords["DELIVERY"]}

    if keywords["INSIDE"] is not None:
        sub = rows[i + 1]
        in_start, out_start = keywords["INSIDE"], keywords["OUTSIDE"]
        boundaries = [v for k, v in keywords.items()
                      if k not in ("INSIDE", "OUTSIDE") and v is not None and v > out_start]
        out_end = min(boundaries) if boundaries else len(sub)

        def find_in(sub_row, lo, hi, label):
            for idx in range(lo, hi):
                if clean(sub_row[idx]).upper() == label:
                    return idx
            return None

        schema["layout"] = "LOC"
        schema["inside_size"] = find_in(sub, in_start, out_start, "SIZE")
        schema["inside_qty"] = find_in(sub, in_start, out_start, "QUANTITY")
        schema["outside_size"] = find_in(sub, out_start, out_end, "SIZE")
        schema["outside_qty"] = find_in(sub, out_start, out_end, "QUANTITY")
        return schema, i + 2

    if keywords["CLASSIFICATION"] is not None:
        schema["layout"] = "CLASS"
        schema["classification"] = keywords["CLASSIFICATION"]
        schema["qty"] = keywords["QUANTITY"]
        return schema, i + 1

    if keywords["SIZE"] is not None:
        schema["layout"] = "SIZE"
        schema["size"] = keywords["SIZE"]
        schema["qty"] = keywords["QUANTITY"]
        return schema, i + 1

    schema["layout"] = "PLAIN"
    schema["qty"] = keywords["QUANTITY"]
    return schema, i + 1


def get(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def convert_sheet(ws_rows, sheet_name, out_rows):
    category = category_for(sheet_name)
    report_date = find_report_date(ws_rows)
    report_month = find_report_month(ws_rows, report_date)

    schema = None
    group_no, group_item, group_price = "", "", None

    i = 0
    while i < len(ws_rows):
        row = ws_rows[i]
        is_header = any(clean(c).upper() == "PRICE" for c in row) and any(clean(c).upper() == "ITEM" for c in row)
        if is_header:
            schema, next_i = parse_header_block(ws_rows, i)
            group_no, group_item, group_price = "", "", None
            i = next_i  # skips the LOC layout's SIZE/QUANTITY sub-header row too
            continue
        i += 1
        if schema is None:
            continue

        item_cell = clean(get(row, schema["item"]))
        no_cell = clean(get(row, schema["no"]))
        price_cell = get(row, schema["price"])

        if item_cell:
            group_item = item_cell
            group_no = no_cell
            group_price = parse_num(price_cell)
        no_val = no_cell if no_cell else group_no
        price_val = parse_num(price_cell) if price_cell not in (None, "") else group_price

        base = dict(Category=category, Date=to_iso_date(report_date, report_month),
                    No=no_val, Item=group_item, Price=price_val if price_val is not None else "")

        opex = parse_num(get(row, schema.get("opex"))) if schema.get("opex") is not None else None

        if schema["layout"] == "LOC":
            in_size = clean(get(row, schema["inside_size"]))
            in_qty = parse_num(get(row, schema["inside_qty"]))
            out_size = clean(get(row, schema["outside_size"]))
            out_qty = parse_num(get(row, schema["outside_qty"]))
            if not (in_size or in_qty or out_size or out_qty):
                continue
            if in_size or in_qty is not None:
                out_rows.append({**base, "Size": in_size, "Location": "INSIDE",
                                  "Classification": "", "Quantity": in_qty if in_qty is not None else "",
                                  "OPEX": opex if opex is not None else "", "Notes": ""})
            if out_size or out_qty is not None:
                out_rows.append({**base, "Size": out_size, "Location": "OUTSIDE",
                                  "Classification": "", "Quantity": out_qty if out_qty is not None else "",
                                  "OPEX": opex if opex is not None else "", "Notes": ""})

        elif schema["layout"] == "CLASS":
            classification = clean(get(row, schema["classification"]))
            qty = parse_num(get(row, schema["qty"]))
            if not (classification or qty is not None):
                continue
            notes = clean(get(row, schema.get("delivery")))
            out_rows.append({**base, "Size": "", "Location": "", "Classification": classification,
                              "Quantity": qty if qty is not None else "",
                              "OPEX": opex if opex is not None else "", "Notes": notes})

        elif schema["layout"] == "SIZE":
            size = clean(get(row, schema["size"]))
            qty = parse_num(get(row, schema["qty"]))
            if not (size or qty is not None):
                continue
            taken = clean(get(row, schema.get("taken_items")))
            out_rows.append({**base, "Size": size, "Location": "", "Classification": "",
                              "Quantity": qty if qty is not None else "",
                              "OPEX": opex if opex is not None else "", "Notes": taken})

        else:  # PLAIN - one row per item, no size/classification breakdown, so no
               # forward-filled continuation rows; a blank Item here means a section
               # label ("USTORE ITEMS", "JUC ITEMS", ...) or spacer row, not data.
            if not item_cell:
                continue
            qty = parse_num(get(row, schema["qty"]))
            notes = "; ".join(clean(get(row, c)) for c in
                               (schema.get("taken_items"), schema.get("delivery")) if c is not None)
            notes = "; ".join(n for n in notes.split("; ") if n)
            out_rows.append({**base, "Size": "", "Location": "", "Classification": "",
                              "Quantity": qty if qty is not None else "",
                              "OPEX": opex if opex is not None else "", "Notes": notes})


def convert(in_path, out_path):
    wb = openpyxl.load_workbook(in_path, read_only=True, data_only=True)
    out_rows = []
    for sheet_name in wb.sheetnames:
        ws_rows = [list(r) for r in wb[sheet_name].iter_rows(values_only=True)]
        convert_sheet(ws_rows, sheet_name, out_rows)
    wb.close()

    fields = ["Category", "Date", "No", "Item", "Size", "Location",
              "Classification", "Price", "Quantity", "OPEX", "Notes"]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(out_rows)
    return len(out_rows)


def main():
    ap = argparse.ArgumentParser(description="Convert the USTore Excel inventory workbook to a tidy long-format CSV.")
    ap.add_argument("file", nargs="?", default=DEFAULT_INPUT, help="Input inventory .xlsx")
    ap.add_argument("-o", "--output", default=DEFAULT_OUTPUT, help="Output CSV path")
    args = ap.parse_args()

    n = convert(args.file, args.output)
    print(f"Wrote {n} rows -> {args.output}")


if __name__ == "__main__":
    main()
