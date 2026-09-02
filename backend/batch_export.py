"""
batch_export.py — CSV and XLSX renderers for the monthly Batch Sales Report.

Shares build_batch_report()'s output with the PDF (batch_pdf.py) and the
on-screen table, so all three export formats and the screen can never show
different numbers for the same month.

Column names match the store's own TBS workbooks (rawdata/*.xlsx: ITEM,
TOTAL QUANTITY, ITEM PRICE, FOR REMITTANCE per supplier block) rather than
inventing new terminology - the point of these exports is to be the same
shape as the sheet Purchasing/Finance already work from.

Units, not pesos, are the SUBTOTAL/TOTAL rows - same rule the PDF and the
screen follow (see batch_pdf.py's docstring): this is an internal counting
document for supplier remittance reference, not an invoice (BIR constraint,
docs/PROMPT_1_FRONTEND.md §1). Item Price and For Remittance appear as
per-line reference data only.
"""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from batch_pdf import long_month

COLUMNS = ["Supplier", "Item", "Quantity", "Item Price (PHP)", "For Remittance (PHP)"]

GOLD_FILL = PatternFill(start_color="F4E4B8", end_color="F4E4B8", fill_type="solid")
INK_FILL = PatternFill(start_color="16140F", end_color="16140F", fill_type="solid")
WHITE_FONT = Font(bold=True, color="FFFFFF")


def _line_rows(report):
    """One row per line item, then a SUBTOTAL row per supplier — the same
    grouping the store's own TBS sheets use per-supplier block."""
    for entry in report:
        for item in entry["items"]:
            yield [
                entry["supplier"],
                item["item_name"],
                item["quantity"],
                item["unit_price_php"] if item["unit_price_php"] is not None else None,
                item["line_total"] if item["line_total"] is not None else None,
            ]
        yield [f"SUBTOTAL — {entry['supplier']}", "", entry["total_units"], None, None]


def render_csv(report, month):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["USTore Monthly Batch Sales Report"])
    writer.writerow([f"Period: {long_month(month)}"])
    writer.writerow([])
    writer.writerow(COLUMNS)
    for row in _line_rows(report):
        writer.writerow(["" if v is None else v for v in row])
    if report:
        writer.writerow([])
        writer.writerow(["GRAND TOTAL UNITS SOLD — ALL SUPPLIERS", "",
                         sum(e["total_units"] for e in report), "", ""])
    # utf-8-sig: the BOM is what makes Excel open this as UTF-8 rather than
    # guessing a legacy codepage and mangling the peso-adjacent text.
    return buf.getvalue().encode("utf-8-sig")


def render_xlsx(report, month):
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Sales"

    ws.append(["USTore Monthly Batch Sales Report"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([f"Period: {long_month(month)}"])
    ws.append([])
    ws.append(COLUMNS)
    for cell in ws[ws.max_row]:
        cell.font = WHITE_FONT
        cell.fill = INK_FILL

    for row in _line_rows(report):
        ws.append(row)
        if str(row[0]).startswith("SUBTOTAL"):
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
                cell.fill = GOLD_FILL

    if report:
        ws.append([])
        ws.append(["GRAND TOTAL UNITS SOLD — ALL SUPPLIERS", "",
                   sum(e["total_units"] for e in report), "", ""])
        for cell in ws[ws.max_row]:
            cell.font = WHITE_FONT
            cell.fill = INK_FILL

    for i, width in enumerate([34, 42, 12, 16, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
