"""
tools/provenance_may2024.py
------------------------------------------------------------------
Reproduces the 88,481 -> 89,232 change in SUM(quantity_sold).

This script REPRODUCES a completed analysis; it does not investigate one.
Every expected value below is an input, re-derived from a from-scratch
rebuild on 2026-08-04. Never edit one to make the script pass.

The delta decomposes into exactly two months and no others:

    2024-05    4,318 -> 4,022    -296     source-sheet change (DSR -> TBS)
    2026-07        0 -> 1,047  +1,047     a genuinely new month
    net                           +751

2026-07 is new data. 2024-05 is NOT a new month - both files cover the
same 23 tally dates. The difference is which sheet the month was read
from:

  * The old converter ("Converter Aug 2024 - May 2026.py") only accepts a
    sheet whose name ends in "- TBS" (is_tbs_month_sheet), and its
    default_files list does not include the May 2024 workbook at all. That
    workbook's tally sheet is named plain "TBS", so it could never have
    been picked up by that path.
  * The old CSV's May 2024 therefore came from the workbook's 23 daily
    "DAILY SALES REPORT" sheets (May 2, May 3, ...), which label items
    WITHOUT a price and carry separate RETAIL PRICE / DISCOUNTED PRICE /
    SPECIAL DISC. PRICE columns, each with its own PCS SOLD.
  * step0_convert_sales_with_zeros.py hardcodes sheet_names = ["TBS"] for
    that workbook, and the TBS sheet folds the price into the item label.

The signature of this is the sticker line. The DSR sheets carry one
unpriced "Sticker" label at 336 units; the TBS sheet splits stickers
across priced SKUs totalling 65 (22 + 24 + 7 + 12).

Run:
    python tools/provenance_may2024.py

Exit 0 = the decomposition holds. Outputs docs/may2024_dsr_vs_tbs.csv
and docs/DATA_PROVENANCE.md.
------------------------------------------------------------------
"""
import csv
import os
import sys
from collections import defaultdict

import openpyxl
import pandas as pd

RAW_CSV = "data/USTore_sales_long_May_Aug2024-May2026.csv"
ZEROFILL_CSV = "data/USTore_sales_long_with_zeros.csv"
WORKBOOK = "drive-download-20260724T120738Z-1-001/2024 5 MAY DSR & TBS.xlsx"
DOCS = "docs"
DIFF_CSV = os.path.join(DOCS, "may2024_dsr_vs_tbs.csv")
NARRATIVE_MD = os.path.join(DOCS, "DATA_PROVENANCE.md")

TBS_SHEET = "TBS"
QTY = "Total Quantity"

# ---- expected values: inputs, not outputs -------------------------
EXP_TBS_DATE_COLUMNS = 23
EXP_TBS_GRAND_TOTAL = 4022
EXP_MONTHS_MOVED = ["2024-05", "2026-07"]
EXP_DELTA_2024_05 = -296
EXP_DELTA_2026_07 = 1047
EXP_NET_DELTA = 751
EXP_DSR_MAY_STICKER = 336
EXP_TBS_STICKER = 65          # 22 + 24 + 7 + 12
# The three DSR price channels sum to exactly the old CSV's May 2024 figure.
# This is the check that closes the argument: it is not "the DSR is roughly
# where 4,318 came from", it is "the DSR is exactly where 4,318 came from".
EXP_DSR_GRAND_TOTAL = 4318
EXP_RAW_MAY_TOTAL = 4318


def month_totals(path):
    df = pd.read_csv(path)
    df["month"] = pd.to_datetime(df["Date"], format="%Y-%m-%d").dt.strftime("%Y-%m")
    return df.groupby("month")[QTY].sum().astype(int)


def read_tbs(wb):
    """TBS sheet: date columns only (the meta columns TOTAL QUANTITY /
    ITEM PRICE / FOR REMITTANCE / REMARKS are excluded), one total per
    priced item label."""
    ws = wb[TBS_SHEET]
    date_cols = [c for c in range(2, ws.max_column + 1)
                 if hasattr(ws.cell(1, c).value, "year")]
    items = defaultdict(float)
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None or str(label).strip() == "":
            continue
        label = str(label).strip()
        if label.upper() == "TOTAL":
            continue
        for c in date_cols:
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                items[label] += float(v)
    return len(date_cols), items


def read_dsr(wb):
    """The 23 daily DAILY SALES REPORT sheets. Quantities live in the
    'PCS SOLD' columns - one under RETAIL PRICE, one under DISCOUNTED
    PRICE, one under SPECIAL DISC. PRICE. Column positions are resolved
    from the header row rather than hardcoded."""
    sheets = [s for s in wb.sheetnames
              if s.strip().upper().startswith("MAY") and s.strip().upper() != TBS_SHEET]
    items = defaultdict(float)
    by_channel = defaultdict(float)

    for sn in sheets:
        ws = wb[sn]
        header_row = None
        for r in range(1, 8):
            if str(ws.cell(r, 1).value or "").strip().upper().startswith("ITEMS"):
                header_row = r
                break
        if header_row is None:
            continue

        # each "PCS SOLD" belongs to the price column most recently seen
        pcs_cols = []
        channel = "RETAIL"
        for c in range(2, ws.max_column + 1):
            h = str(ws.cell(header_row, c).value or "").strip().upper()
            if "PRICE" in h and "PCS" not in h:
                channel = h.replace(" PRICE", "").strip()
            elif h.startswith("PCS SOLD"):
                pcs_cols.append((c, channel))

        for r in range(header_row + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if label is None or str(label).strip() == "":
                continue
            label = str(label).strip()
            if label.upper() in ("TOTAL", "ITEMS"):
                continue
            for c, ch in pcs_cols:
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    items[label] += float(v)
                    by_channel[ch] += float(v)
    return len(sheets), items, by_channel


def sticker_total(items):
    return int(sum(v for k, v in items.items() if "STICKER" in k.upper()))


def main():
    os.makedirs(DOCS, exist_ok=True)
    failures = []

    def expect(label, actual, expected):
        ok = actual == expected
        print("[%s] %-34s %r%s" % ("PASS" if ok else "FAIL", label, actual,
                                   "" if ok else "   != expected %r" % (expected,)))
        if not ok:
            failures.append((label, actual, expected))

    # ---- 1. month decomposition -----------------------------------
    raw, zf = month_totals(RAW_CSV), month_totals(ZEROFILL_CSV)
    cmp = pd.DataFrame({"raw": raw, "zerofill": zf}).fillna(0).astype(int)
    cmp["delta"] = cmp["zerofill"] - cmp["raw"]
    moved = cmp.index[cmp["delta"] != 0].tolist()

    expect("months with non-zero delta", moved, EXP_MONTHS_MOVED)
    expect("delta 2024-05", int(cmp.loc["2024-05", "delta"]), EXP_DELTA_2024_05)
    expect("delta 2026-07", int(cmp.loc["2026-07", "delta"]), EXP_DELTA_2026_07)
    expect("net delta", int(cmp["delta"].sum()), EXP_NET_DELTA)

    # ---- 2. the TBS sheet is faithful -----------------------------
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    n_date_cols, tbs_items = read_tbs(wb)
    expect("TBS date columns", n_date_cols, EXP_TBS_DATE_COLUMNS)
    expect("TBS grand total", int(sum(tbs_items.values())), EXP_TBS_GRAND_TOTAL)

    # ---- 3. the DSR sheets are the old provenance -----------------
    n_dsr, dsr_items, by_channel = read_dsr(wb)
    expect("DSR daily sheets", n_dsr, EXP_TBS_DATE_COLUMNS)   # same 23 tally dates
    expect("DSR 'Sticker' units", int(dsr_items.get("Sticker", 0)), EXP_DSR_MAY_STICKER)
    expect("TBS sticker units", sticker_total(tbs_items), EXP_TBS_STICKER)
    expect("DSR grand total", int(sum(by_channel.values())), EXP_DSR_GRAND_TOTAL)
    expect("old CSV May 2024 total", int(cmp.loc["2024-05", "raw"]), EXP_RAW_MAY_TOTAL)

    # ---- 4. item-level diff ---------------------------------------
    raw_df = pd.read_csv(RAW_CSV)
    raw_df = raw_df[raw_df["Date"].str.startswith("2024-05")]
    dsr_csv = raw_df.groupby("Item")[QTY].sum().astype(int).to_dict()

    labels = sorted(set(dsr_csv) | set(k for k, v in tbs_items.items() if v))
    with open(DIFF_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["item_label", "dsr_units", "tbs_units", "delta", "present_in"])
        for lab in labels:
            d = int(dsr_csv.get(lab, 0))
            t = int(tbs_items.get(lab, 0))
            where = "both" if (d and t) else ("dsr_only" if d else "tbs_only")
            w.writerow([lab, d, t, t - d, where])

    print("\nWrote %s (%d labels)" % (DIFF_CSV, len(labels)))
    print("DSR channel split: %s" % {k: int(v) for k, v in sorted(by_channel.items())})

    write_narrative(cmp, n_date_cols, tbs_items, dsr_items, by_channel, labels, dsr_csv)
    print("Wrote %s" % NARRATIVE_MD)

    if failures:
        print("\nFAILED: %d check(s). This is a finding - record it under "
              "'Gate failures' in docs/CHANGES_tyrone.md. Do not edit the expected values."
              % len(failures))
        for lab, a, e in failures:
            print("  %s: actual=%r expected=%r" % (lab, a, e))
        return 1
    print("\nAll provenance checks passed.")
    return 0


def write_narrative(cmp, n_date_cols, tbs_items, dsr_items, by_channel, labels, dsr_csv):
    dsr_only = [l for l in labels if dsr_csv.get(l) and not tbs_items.get(l)]
    tbs_only = [l for l in labels if tbs_items.get(l) and not dsr_csv.get(l)]
    ch = {k: int(v) for k, v in sorted(by_channel.items())}

    lines = []
    a = lines.append
    a("# Data provenance: why SUM(quantity_sold) is 89,232 and not 88,481")
    a("")
    a("Generated by `tools/provenance_may2024.py`. Every figure here is asserted by")
    a("that script and by `verify_data.py`; neither is a hand-typed number.")
    a("")
    a("## The decomposition")
    a("")
    a("`USTore_sales_long_with_zeros.csv` minus `USTore_sales_long_May_Aug2024-May2026.csv`,")
    a("by month. Exactly two months move; every other month totals the same to the unit.")
    a("")
    a("| Month | Old file | Zero-filled | Delta |")
    a("|---|---:|---:|---:|")
    for m in cmp.index[cmp["delta"] != 0]:
        a("| %s | %d | %d | **%+d** |" % (m, cmp.loc[m, "raw"], cmp.loc[m, "zerofill"],
                                          cmp.loc[m, "delta"]))
    a("| **Net** | | | **%+d** |" % int(cmp["delta"].sum()))
    a("")
    a("Of the %d months present, %d are unchanged." %
      (len(cmp), int((cmp["delta"] == 0).sum())))
    a("")
    a("## 2026-07 is a new month")
    a("")
    a("The old combined CSV ends at 2026-06-30. The zero-fill rebuild picked up a")
    a("July 2026 sheet the old file never had: +1,047 units. Nothing subtle here.")
    a("")
    a("## 2024-05 is a re-sourcing, not new data")
    a("")
    a("This is the part worth reading. **Both files cover the same %d tally dates**" % n_date_cols)
    a("in May 2024 - the month did not gain or lose a single observation date. What")
    a("changed is which sheet of `2024 5 MAY DSR & TBS.xlsx` the month was read from.")
    a("")
    a("- The old converter (`Converter Aug 2024 - May 2026.py`) accepts a sheet only if")
    a("  its name ends in `- TBS` (`is_tbs_month_sheet`). The May 2024 workbook's tally")
    a("  sheet is named plain `TBS`, and the workbook is not in that converter's")
    a("  `default_files` list either. It could not have supplied May 2024 by that path.")
    a("- The old CSV's May 2024 therefore came from the workbook's %d daily" % n_date_cols)
    a("  `DAILY SALES REPORT` sheets (`May 2`, `May 3`, ...). Those label items")
    a("  **without a price** and carry separate `RETAIL PRICE`, `DISCOUNTED PRICE` and")
    a("  `SPECIAL DISC. PRICE` columns, each with its own `PCS SOLD`.")
    a("- `step0_convert_sales_with_zeros.py` hardcodes `sheet_names = [\"TBS\"]` for this")
    a("  workbook, and the TBS sheet **folds the price into the item label**.")
    a("")
    a("### The sticker signature")
    a("")
    a("The clearest fingerprint of the two labelling conventions:")
    a("")
    a("| Source | Labels | Units |")
    a("|---|---|---:|")
    a("| DSR daily sheets | one unpriced `Sticker` | **%d** |" % int(dsr_items.get("Sticker", 0)))
    a("| TBS sheet | `Sticker @ 115` + `Sticker @20` + `Long Sticker` + `Sticker Pack` | **%d** |"
      % sticker_total(tbs_items))
    a("")
    a("%d DSR labels have no TBS counterpart and %d TBS labels have no DSR counterpart;"
      % (len(dsr_only), len(tbs_only)))
    a("the item-level detail is in `may2024_dsr_vs_tbs.csv`.")
    a("")
    a("### Where the 296 units sit")
    a("")
    a("The DSR sheets split their quantities across three price channels:")
    a("")
    a("| Channel | Units |")
    a("|---|---:|")
    for k, v in ch.items():
        a("| %s | %d |" % (k, v))
    a("| **DSR total** | **%d** |" % sum(ch.values()))
    a("| **TBS total** | **%d** |" % int(sum(tbs_items.values())))
    a("")
    a("The DSR channels sum to **%d**, which is exactly the old CSV's May 2024 figure."
      % sum(ch.values()))
    a("That equality is asserted, and it is what closes the argument: the old series")
    a("did not merely resemble the DSR sheets, it *was* the DSR sheets.")
    a("")
    a("The TBS sheet has no equivalent breakdown - it records one number per priced")
    a("SKU per date. That the DSR carries %d units outside its retail channel is the"
      % (sum(ch.values()) - ch.get("RETAIL", 0)))
    a("leading explanation for the gap, but confirming it is a question for USTore")
    a("staff, not for this script. It is logged as **B7** in the deferred-decision")
    a("register and is deliberately left open.")
    a("")
    a("## What this means for Chapter 4")
    a("")
    a("The manuscript's **88,481** was computed on a mixed-provenance series: May 2024")
    a("read from daily DSR sheets, every other month read from TBS tally sheets.")
    a("**89,232** is the single-provenance figure - every month read from the TBS")
    a("tally sheets, plus the July 2026 month the old file predated.")
    a("")
    a("89,232 is the correct invariant. It is asserted in `verify_data.py`,")
    a("`tools/assert_invariants.py` and this script. See Divergence Register #18.")
    a("")

    with open(NARRATIVE_MD, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    sys.exit(main())
